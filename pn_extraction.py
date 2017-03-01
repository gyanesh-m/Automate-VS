import os
from xlrd import *
from openpyxl import Workbook,load_workbook
from extraction import Extraction
class Part_number_extraction(Extraction):
    def __init__(self,alias_data2,standard_list,w_f,w_op,size,convert_i_2_nb,p1,p2,convert_nb_2_i):##update restwhere
        self.standard_list=standard_list
        self.w_f=w_f
        self.w_op=w_op
        self.size=size
        self.i2n=convert_i_2_nb
        self.p1=p1
        self.p2=p2
        self.alias_data2=alias_data2
        self.size_dict=convert_nb_2_i
        if self.p1=='YES':
            self.main()
        if self.p2 =='YES':
            self.main2()
        print "Started part number extraction"

    def sap_normalised(self,col_name,row_n,conversion_d,size_col,r_col,av_r):
        norm={}
        val=0
        for i in range(len(row_n)):
            if i!=size_col and row_n[i]!='':
                if i==r_col:
                    rat=''
                    try:
                        for val in str(row_n[i]):
                            try:
                                if type(int(val))==int or type(float(val))==float:
                                    rat+=str(val)

                            except:
                                    pass
                        norm[i]=conversion_d[rat]
                    except:
                        if rat=='':
                            rat='0'
                        for rating in av_r:

                            if rating>=float(rat):
                                val=rating
                                break
                        norm[i]=conversion_d[val]
                else:
                    try:
                        norm[i]=conversion_d[row_n[i]]
                    except:
                        if row_n[i] in ['BW','FL','SW','SCR']:
                            if col_name=='V_FH':
                                val='SW'
                            elif col_name=='V_CS':
                                val='BW'
                            elif col_name =='V_TAHH':
                                val='BW'
                            norm[i]=conversion_d[val]
            else:
                if row_n[i] !='3':
                    if self.size=='IN':
                        norm[i]=row_n[i]+'\"'
                    else:
                        try:
                            norm[i]=self.size_dict[float(row_n[i])]
                        except:
                            pass
                else:
                
                    norm[i]=row_n[i]
        return norm
    def standard(self,nons):
        k=''
        for i in nons:
            if i in '0123456789':
                k+=i
        return k
    def assign_partn(self,records,p_d,row_x,sheet_f):
            temp={}
            for i in records:   
                    temp[float(p_d[i[0]])]=i[0]
               
            try:
                del temp[0.0]
            except:
                pass
            maxx=max(temp.keys())
            minn=min(temp.keys())
            sheet_f.cell(row=row_x+1,column=sheet_f.max_column-3).value=temp[maxx]
            sheet_f.cell(row=row_x+1,column=sheet_f.max_column-2).value=maxx
            sheet_f.cell(row=row_x+1,column=sheet_f.max_column-1).value=temp[minn]
            sheet_f.cell(row=row_x+1,column=sheet_f.max_column).value=minn
    def format_n_filter(self,length,alias_data,sheet_row,data_column,size,i2n):
        temp_l=[]
        temp=''
        for data in data_column:
                try:
                    cap=sheet_row[data[1]]
                except Exception as e:
                    print e
                    continue
               
                if ('-' not in str(cap) or '/' not in str(cap) )and data[0]!='SIZE':
                    cap=Extraction.capitalised(self,str(cap))
                for key in alias_data:
                    for alias in alias_data[key]:
                        if cap in alias_data[key][alias]:
                            temp_l.append(alias)
                if data[0]=='SIZE':
                   if size=='IN':
                        cap+='\"'
                        temp_l.append(int(self.i2n[cap]))
                   else:
                        if cap!='':
                            try:
                                temp_l.append(int(cap))
                            except Exception as e:
                                print e
                                pass
                rat=''
                if data[0]=='RATING':
                    for val in str(cap):
                        try:
                            if type(int(val))==int or type(float(val))==float:
                                rat+=str(val)
                        except:
                                pass
                        
                    temp_l.append(int(rat))

        for i in range(length):
            try:
                temp+='-'+str(temp_l[i])
            except:
                temp+='-'
        temp=temp[1:]+'-'
        return temp
    def main2(self):
        record=''
        vcode,vprice='',''
        db_w=open_workbook(os.getcwd()+'/'+'Offline database.xls')
        db_s=db_w.sheet_by_name('main sheet')
        sheet_op=self.w_op.get_sheet_by_name('final')
        sheet_f=self.w_f.sheet_by_name('final')
        max_c=sheet_op.max_column
        des,valc,price=0,8,14
        db_dict={}
        header_f={}
        for j in range(sheet_f.ncols):
            header_f[sheet_f.cell(0,j).value]=j
        for i in range(1,db_s.nrows):
            db_dict[db_s.cell(i,des).value]=(db_s.cell(i,valc).value,db_s.cell(i,price).value)
        order_db=['TYPE','SIZE','RATING','MATERIAL','END','OPERATION']
        order_d=()
        for val in order_db:
            for val2 in header_f:
                if val in val2.replace(' ','') or val2.replace(' ','') in val and val2 !='':
                    order_d+=(val,header_f[val2]),
        key_db=db_dict.keys()
        key_db.sort()
        sheet_op.cell(row=1,column=max_c+1).value='CODE'
        sheet_op.cell(row=1,column=max_c+2).value='PRICE'
        for row_x in range(1,sheet_f.nrows):
            try:
                record=self.format_n_filter(len(order_d),self.alias_data2,sheet_f.row_values(row_x),order_d,self.size,self.i2n)
                vcode,vprice=db_dict[record]
                
                sheet_op.cell(row=row_x+1,column=max_c+1).value=vcode
                sheet_op.cell(row=row_x+1,column=max_c+2).value=vprice
            except Exception as e:
                if record!='':
                    li=record.split('-')
                    li1,li2=li[0],li[3]
                    for i in range(1,2):
                        li1+='-'+li[i]
                    for i in range(4,len(li)):
                        li2+='-'+li[i]
                    rat=''
                    for val in li[2]:
                        try:
                            if type(int(val))==int or type(float(val))==float:
                                rat+=str(val)
                        except:
                                pass
                    if rat !='':
                        rat=int(rat)
                    for valx in key_db:
                        if li1 in valx and li2 in valx:
                            rat_a=int(valx.split('-')[2])
                            try:
                                if rat_a > rat:
                                    record=li1+'-'+str(rat_a)+'-'+li2
                                    vcode,vprice=db_dict[record]
                                    sheet_op.cell(row=row_x+1,column=max_c+1).value=vcode
                                    sheet_op.cell(row=row_x+1,column=max_c+2).value=vprice
                                    break
                            except:
                                continue
        
        self.w_op.save(os.getcwd()+'/'+'Output.xls')
    def main(self):
        con=cx_Oracle.connect('username','password','10.1.40.24:1521/webos')
        fm=con.cursor()
        fh=con.cursor()
        cs=con.cursor()
        ta=con.cursor()
        cs.arraysize=1000
        ta.arraysize=1000
        fm.arraysize=1000
        fh.arraysize=1000
        fm.prepare('SELECT * FROM itsstest.VAL21_SAP_FM WHERE SVSIZE=:VALVESIZE and SVENDCONN=:ENDCONNECTION AND SVOPER=:OPERATION and SVTYPE=:TYPE AND SVMATL=:MATERIAL AND SVRATING=:RATING')
        fh.prepare('SELECT * FROM itsstest.VAL23_SAP_FH WHERE SVSIZE=:VALVESIZE and SVENDCONN=:ENDCONNECTION AND SVOPER=:OPERATION and SVTYPE=:TYPE AND SVMATL=:MATERIAL AND SVRATING=:RATING')
        cs.prepare('SELECT * FROM itsstest.VAL24_SAP_CS WHERE SVSIZE=:VALVESIZE and SVENDCONN=:ENDCONNECTION AND SVOPER=:OPERATION and SVTYPE=:TYPE AND SVMATL=:MATERIAL AND SVRATING=:RATING')
        ta.prepare('SELECT * FROM itsstest.VAL22_SAP_TA WHERE SVSIZE=:VALVESIZE and SVENDCONN=:ENDCONNECTION AND SVOPER=:OPERATION and SVTYPE=:TYPE AND SVMATL=:MATERIAL AND SVRATING=:RATING')
        sheet_f=self.w_f.sheet_by_name('final')#xlrd
        bhel_r=self.standard_list.sheet_by_name('bhel_rating')
        available_r={}
        for col in range(1,bhel_r.ncols):
            temp=[]
            for row in range(1,bhel_r.nrows):
                if bhel_r.cell(row,col).value!='':
                    temp.append(bhel_r.cell(row,0).value)
            available_r[bhel_r.cell(0,col).value]=temp
        sap_table=self.standard_list.sheet_by_name('table')
        sap_dict={}
        sap_d={}
        for col_t in range(1,sap_table.ncols):
            sap_d={}
            for row_t in range(1,sap_table.nrows):
                if sap_table.cell(row_t,0).value!='' and sap_table.cell(row_t,col_t).value!='':
                    sap_d[sap_table.cell(row_t,0).value]=sap_table.cell(row_t,col_t).value
            sap_dict[sap_table.cell(0,col_t).value]=sap_d
        header_f={}
        for j in range(sheet_f.ncols):
            header_f[sheet_f.cell(0,j).value]=j
        dd=['V_CS','V_FH','V_TAHH','V_FM']
        price_d={}
        for val in dd:
            wb=open_workbook(os.getcwd()+'/'+val+".xls")
            ws1=wb.sheet_by_index(0)
            mater_no=2
            s_price=6
            p_d={}
            for i in range(1,ws1.nrows):
                p_d[ws1.cell(i,mater_no).value]=ws1.cell(i,s_price).value
            price_d[val]=p_d
        ##final_s-size-rating-vtype-end conn-operation-material
        rat=0
        #type size material class operation end connection
        order_l=['OPERATION','ENDCONNECTION','TYPE','RATING','MATERIAL','VALVESIZE']
        order_d={}
        sap={}
        records=[]
        for val in order_l:
            for val2 in header_f:
                if val in val2.replace(' ','') or val2.replace(' ','') in val and val2 !='':
                    order_d[val]=header_f[val2]
        sheet_op=self.w_op.get_sheet_by_name('final')
        sheet_op.cell(row=1,column=sheet_f.ncols+1).value='Max Part Number'
        sheet_op.cell(row=1,column=sheet_f.ncols+2).value='Max Price'
        sheet_op.cell(row=1,column=sheet_f.ncols+3).value='Min Part Number'
        sheet_op.cell(row=1,column=sheet_f.ncols+4).value='Min Price'
        f=open(os.getcwd()+'/'+'test.txt','w+')
        for row in range(1,sheet_f.nrows):
                normalised=[]
                size=str(sheet_f.cell(row,header_f['VALVE SIZE']).value).replace('\\','/').split('-')
                if '/' in str(size) and len(size)>1:
                    a,b=size
                    b=b.replace('/','')
                    size=int(a)+float(b[0])/float(b[1])
                elif '/' in str(size):
                    b,=size
                    b=b.replace('/','')
                    size=float(b[0])/float(b[1])
                elif 'X' not in str(size) and '' not in size:
                    b,=size
                    size=float(b)
                rat=''
                for val in str(sheet_f.cell(row,header_f['VALVE RATING']).value):
                    try:
                        if type(int(val))==int or type(float(val))==float:
                            rat+=str(val)
                    except:
                            pass
                if rat=='':
                    p=1
                else:
                    sap={}
                    try:
                            if float(rat)==800:
                                normalised=self.sap_normalised('V_FM',sheet_f.row_values(row),sap_dict['V_FM'],header_f['VALVE SIZE'],header_f['VALVE RATING'],available_r['V_FM'])
                                for data in order_l:
                                    sap[data]=normalised[order_d[data]]
                                fm.execute(None,**sap)
                                records=fm.fetchall()
                                self.assign_partn(records,price_d['V_FM'],row,sheet_op)
                            elif float(rat)<=900:
                                normalised=self.sap_normalised('V_CS',sheet_f.row_values(row),sap_dict['V_CS'],header_f['VALVE SIZE'],header_f['VALVE RATING'],available_r['V_CS'])
                                for data in order_l:
                                    sap[data]=normalised[order_d[data]]
                                cs.execute(None,**sap)
                                records=cs.fetchall()
                                self.assign_partn(records,price_d['V_CS'],row,sheet_op)
                            elif size<=2:
                                normalised=self.sap_normalised('V_FH',sheet_f.row_values(row),sap_dict['V_FH'],header_f['VALVE SIZE'],header_f['VALVE RATING'],available_r['V_FH'])            #V_FH,V_FM
                                for data in order_l:
                                    sap[data]=normalised[order_d[data]]
                                fh.execute(None,**sap)
                                records=fh.fetchall()
                                self.assign_partn(records,price_d['V_FH'],row,sheet_op)
                            elif size>2:
                                normalised=self.sap_normalised('V_TAHH',sheet_f.row_values(row),sap_dict['V_TAHH'],header_f['VALVE SIZE'],header_f['VALVE RATING'],available_r['V_TAHH'])
                                for data in order_l:
                                    sap[data]=normalised[order_d[data]]
                                ta.execute(None,**sap)
                                records=ta.fetchall()
                                self.assign_partn(records,price_d['V_TAHH'],row,sheet_op)
                    except Exception as e:
                        f.write(str(e)+'---'+'\n'+str(records)+'\n')
                        try:
                            normalised=self.sap_normalised('V_CS',sheet_f.row_values(row),sap_dict['V_CS'],header_f['VALVE SIZE'],header_f['VALVE RATING'],available_r['V_CS'])
                            for data in order_l:
                                sap[data]=normalised[order_d[data]]
                            cs.execute(None,**sap)
                            records=cs.fetchall()
                            self.assign_partn(records,price_d['V_CS'],row,sheet_op)
                        except Exception as e1:
                            f.write(str(e1)+'---'+'\n'+str(records)+'\n')                    
                            try:
                                normalised=self.sap_normalised('V_FM',sheet_f.row_values(row),sap_dict['V_FM'],header_f['VALVE SIZE'],header_f['VALVE RATING'],available_r['V_FM'])
                                for data in order_l:
                                    sap[data]=normalised[order_d[data]]
                                fm.execute(None,**sap)
                                records=fm.fetchall()
                                self.assign_partn(records,price_d['V_FM'],row,sheet_op)


                            except Exception as e2:
                                f.write(str(e2)+'---'+'\n'+str(records)+'\n')
                                try:

                                        normalised=self.sap_normalised('V_FH',sheet_f.row_values(row),sap_dict['V_FH'],header_f['VALVE SIZE'],header_f['VALVE RATING'],available_r['V_FH'])            #V_FH,V_FM
                                        for data in order_l:
                                            sap[data]=normalised[order_d[data]]
                                        fh.execute(None,**sap)
                                        records=fh.fetchall()
                                        self.assign_partn(records,price_d['V_FH'],row,sheet_op)
                                except Exception as e3:
                                    f.write(str(e3)+'---'+'\n'+str(records)+'\n')
                                    try:
                                        normalised=self.sap_normalised('V_TAHH',sheet_f.row_values(row),sap_dict['V_TAHH'],header_f['VALVE SIZE'],header_f['VALVE RATING'],available_r['V_TAHH'])
                                        for data in order_l:
                                            sap[data]=normalised[order_d[data]]
                                        ta.execute(None,**sap)
                                        records=ta.fetchall()
                                        self.assign_partn(records,price_d['V_TAHH'],row,sheet_op)
                                    except Exception as e4:
                                        f.write(str(e4)+'---'+'\n'+str(records)+'\n')
                                        pass
                    print 'total'+str(len(records))
        self.w_op.save(os.getcwd()+'/'+'Output.xls')
        f.close()
