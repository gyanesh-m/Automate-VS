import os
from xlrd import * 
from suds.client import Client
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl import Workbook,load_workbook
#import extraction
from filter_ import Filter
from pn_extraction import Part_number_extraction
from extraction import Extraction

class Base(Extraction,Filter,Part_number_extraction):
    ##inside function of openfilename
    def __init__(self,filename,rating_value,part_value,part_value2,material_value,size_value,vp,vs,a105,f22,f91,f92,wcb,wcc,wc9,c12a,c12ab):
        self.rating_value=rating_value
        self.part_value=part_value
        self.part_value2=part_value2
        self.material_value=material_value
        self.size_value=size_value
        self.fname=filename
        self.vp=vp
        self.vs=vs
        self.a105=a105
        self.f22=f22
        self.f91=f91
        self.f92=f92
        self.wcb=wcb
        self.wcc=wcc
        self.wc9=wc9
        self.c12a=c12a
        self.c12ab=c12ab
        self.main_code(self.fname)
        print 'inside base'
    def main_code(self,fname):
        wb=open_workbook(fname,formatting_info=True)
        standard_list=open_workbook(os.getcwd()+'/Standard Values.xlsx')
        standard_table=Extraction.extract_data(self,standard_list)
        ptm_data=Extraction.extract_data(self,standard_list,'alias_col')
        font=wb.font_list
        sheet_name=[]
        cordinates=Extraction.search_titles(self,wb,standard_table)
        cordinates_row=Extraction.extract_rows(self,cordinates)
        #below code finds the max column
        max_col=1
        max_col_sheet=''
        row_x=0
        for sheetname in cordinates.keys():
            for value in cordinates[sheetname]:
                x,y=value
                if y>max_col and y!='':
                    max_col=y
                    max_col_sheet=sheetname
                    row_x=x
        sheet_m=wb.sheet_by_name(max_col_sheet)
        header={}
        cordinates_dict=Extraction.get_cordinates(self,cordinates_row,wb)
        work=Workbook()
        w_sheet=work.active
        w_sheet.title='Standard Sheet'
        counter=0
        heading_count=0
        x=row_x
        for col in range(sheet_m.ncols):
            if col<sheet_m.ncols-1 and sheet_m.cell(x,col).value!='' and sheet_m.cell(x,col+1).value=='' and sheet_m.cell(x+1,col).value!='':
                w_sheet.cell(row=counter+1,column=col+1).value=sheet_m.cell(x+1,col).value
                header[sheet_m.cell(x+1,col).value]=col
            elif col<sheet_m.ncols-1 and sheet_m.cell(x,col).value=='' and x<sheet_m.nrows-1 :
                if sheet_m.cell(x,col+1).value!='' and sheet_m.cell(x+1,col).value!='':
                    w_sheet.cell(row=counter+1,column=col+1).value=sheet_m.cell(x+1,col).value
                    header[sheet_m.cell(x+1,col).value]=col
                elif sheet_m.cell(x,col+1).value=='' and sheet_m.cell(x+1,col).value!='':
                    w_sheet.cell(row=counter+1,column=col+1).value=sheet_m.cell(x+1,col).value
                    header[sheet_m.cell(x+1,col).value]=col
                elif sheet_m.cell(x,col+1).value!='' and sheet_m.cell(x-1,col).value!='':
                    w_sheet.cell(row=counter+1,column=col+1).value=sheet_m.cell(x-1,col).value
                    header[sheet_m.cell(x-1,col).value]=col
                elif sheet_m.cell(x,col+1).value=='' and sheet_m.cell(x-1,col).value!='':
                    w_sheet.cell(row=counter+1,column=col+1).value=sheet_m.cell(x-1,col).value
                    header[sheet_m.cell(x-1,col).value]=col
            else:
                if sheet_m.cell(x,col).value!='':
                    w_sheet.cell(row=counter+1,column=col+1).value=sheet_m.cell(x,col).value
                    header[sheet_m.cell(row_x,col).value]=col
        counter+=1
        for sheet in wb.sheets():
            if sheet.visibility==0:
                col_n=None
                second_max=0
                print sheet.name
                counting=1
                order_row=(cordinates_dict[sheet.name].keys())
                order_row.sort()
                for tempp in range(len(order_row)-1,-1,-1):
                    if order_row[tempp]<sheet.nrows:
                        second_max=order_row[tempp]
                        break
                for row_n in order_row:#row_n is row in increasing form
                    row_count=row_n
                    if row_count<second_max:#counting is the number of rows in a sheet
                        while row_count<order_row[counting]:
                                for column_name in cordinates_dict[sheet.name][row_n]:
                                        cell_formatting=wb.xf_list[sheet.cell_xf_index(row_count,column_name[1])]
                                        # tests for the presence of striked out cells and skips them
                                        check2=font[cell_formatting.font_index].struck_out!=1
                                        if not check2:
                                            if sheet.cell(row_count,column_name[1]).ctype==XL_CELL_BLANK or sheet.cell(row_count,column_name[1]).ctype==XL_CELL_EMPTY:
                                                check2=True

                                        if check2:
                                            check=Filter.non_zero_row(self,sheet.row_values(row_count),cordinates_dict[sheet.name][row_n],ptm_data['QUANTITY'])
                                            if check:
                                                try:
                                                    w_sheet.cell(row=counter+1,column=header[column_name[0]]+1).value=sheet.cell(row_count,column_name[1]).value
                                                except Exception as e:
                                                    cap1=Extraction.capitalised(self,column_name[0])
                                                    for key in header.keys():
                                                        cap2=Extraction.capitalised(self,key)
                                                        for tempi in ptm_data:
                                                            for j in ptm_data[tempi]:

                                                                if (cap1 in j or j in cap1 )and (j in cap2 or tempi in cap2) :
                                                                    col_n=key
                                                                    break
                                                            if col_n==key:
                                                                break
                                                        if col_n==key:
                                                            break
                                                    w_sheet.cell(row=counter+1,column=header[key]+1).value=sheet.cell(row_count,column_name[1]).value
                                            else:
                                                break
                                        else:
                                            break
                                row_count+=1
                                if check2 and check:
                                    counter+=1
                    elif row_count ==second_max:
                        while row_count < sheet.nrows:
                            for column_name in cordinates_dict[sheet.name][row_n]:
                                cell_formatting=wb.xf_list[sheet.cell_xf_index(row_count,column_name[1])]
                                check2=font[cell_formatting.font_index].struck_out!=1
                                if not check2:
                                    if sheet.cell(row_count,column_name[1]).ctype==XL_CELL_BLANK or sheet.cell(row_count,column_name[1]).ctype==XL_CELL_EMPTY:
                                        check2=True
                                if check2:

                                    check=Filter.non_zero_row(self,sheet.row_values(row_count),cordinates_dict[sheet.name][row_n],ptm_data['QUANTITY'])
                                    if check:
                                        try:
                                            w_sheet.cell(row=counter+1,column=header[column_name[0]]+1).value=sheet.cell(row_count,column_name[1]).value
                                        except Exception as e:
                                            cap1=Extraction.capitalised(self,column_name[0])
                                            for key in header.keys():
                                                cap2=Extraction.capitalised(self,key)
                                                if cap1==cap2:
                                                    break
                                                else:
                                                    for i in ptm_data:
                                                        for j in ptm_data[i]:
                                                            if (cap1 in j or j in cap1 )and (j in cap2 or i in cap2) :
                                                                col_n=key
                                                                break
                                                        if col_n==key:
                                                            break
                                                    if col_n==key:
                                                        break
                                            w_sheet.cell(row=counter+1,column=header[key]+1).value=sheet.cell(row_count,column_name[1]).value
                                    else:
                                        break
                                else:
                                    break
                            if check2 and check:
                                counter+=1
                            row_count+=1
                        counting+=1
                       
                    else:
                        continue
                    counting+=1
        print 'writing done'
        print sheet.name
        work.save(os.getcwd()+'/'+'Output.xls')
        wb2=open_workbook(os.getcwd()+'/'+'Output.xls' )
        sheet=wb2.sheet_by_index(0)
        data_column={}
        for vtemp in header:
            stamp=0
            for key in ptm_data:
                for value in ptm_data[key]:
                    cap=str(Extraction.capitalised(self,vtemp))
                    cap2=str(Extraction.capitalised(self,value))
                    if cap!='' and cap == cap2 or cap2 in cap or cap in cap2:
                        data_column[key]=header[vtemp]
                        stamp=1
                        break
                if stamp==1:
                    break
        
        '''below code is used to construct a nested dictionary of aliases for material, operation, valve type, etc'''
        alias_s=standard_list.sheet_by_name('alias')
        alias_db=standard_list.sheet_by_name('db alias')
        size_con=standard_list.sheet_by_name('conversion')
        convert_nb_2_i={}
        convert_i_2_nb={}
        for row in range(size_con.nrows):
           word=''
           for val in str(size_con.cell(row,1).value):
                word+=val
           convert_nb_2_i[size_con.cell(row,0).value]=word
           convert_i_2_nb[word]=size_con.cell(row,0).value
        if self.size_value=='IN':
         conversion_data=convert_nb_2_i
         for row in range(1,sheet.nrows):
            d_value=''
            value_string=str(sheet.cell(row,data_column['VALVE SIZE']).value).lstrip()
            for i in str(value_string):
                if i in '0123456789':
                    d_value+=i
                else:
                    break
            try:
                w_sheet.cell(row=row+1,column=data_column['VALVE SIZE']+1).value=conversion_data[int(d_value)].strip("\"")
            except:
                pass
        elif self.size_value=='NB':
            conversion_data=convert_i_2_nb
            for row in range(1,sheet.nrows):
                d_value=''
                value_string=str(sheet.cell(row,data_column['VALVE SIZE']).value).strip().strip("\"")
                for i in str(value_string):
                    if i in '0123456789':
                        d_value+=i
                    else:
                        break
                try:
                    w_sheet.cell(row=row+1,column=data_column['VALVE SIZE']+1).value=conversion_data[int(d_value)]
                except:

                    pass
        alias_data={}
        for col in range(alias_s.ncols):
            alias_d={}
            temp=[]
            for row in range(1,alias_s.nrows):
                if alias_s.cell(row,col).value!='':
                    if alias_s.cell(row,col).value!='-' and  alias_s.cell(row,col).value!='--':
                        temp.append(str(Extraction.capitalised(self,alias_s.cell(row,col).value)))
                    else:
                       temp.append(str(alias_s.cell(row,col).value))
                elif alias_s.cell(row,col).value=='' and len(temp)>=1:
                    alias_d[temp[0]]=temp
                    temp=[]
            alias_data[alias_s.cell(0,col).value]=alias_d
        alias_data2={}
        for col in range(alias_db.ncols):
            alias_d={}
            temp=[]
            for row in range(1,alias_db.nrows):
                if alias_db.cell(row,col).value!='':
                    if alias_db.cell(row,col).value!='-' and  alias_db.cell(row,col).value!='--':
                        temp.append(str(Extraction.capitalised(self,alias_db.cell(row,col).value)))
                    else:
                       temp.append(str(alias_db.cell(row,col).value))
                elif alias_db.cell(row,col).value=='' and len(temp)>=1:
                    alias_d[temp[0]]=temp
                    temp=[]
            alias_data2[alias_db.cell(0,col).value]=alias_d
        C=0

        '''Below code is used to standardise the material name given in the original excel sheetusing alias excel sheet'''
        work.save(os.getcwd()+'/'+'Output.xls')
        wb2=open_workbook(os.getcwd()+'/'+'Output.xls' )
        Filter.formatting_data(self,convert_nb_2_i,self.size_value,wb2,work,data_column,alias_data,self.material_value,self.a105,self.f22,self.f91,self.f92,self.wcb,self.wcc,self.wc9,self.c12a,self.c12ab)
        work.save(os.getcwd()+'/'+'Output.xls')
        wb2=open_workbook(os.getcwd()+'/'+'Output.xls' )
        sheet=wb2.sheet_by_index(0)
        if self.rating_value=='YES':
            cliet=Client('enter ip address for the soap service')
            r=[]
            w_sheet.cell(row=1,column=sheet.ncols+1).value="RATING(S)"
            data_column['VALVE RATING']=sheet.ncols
            print ("FETCHING RATING FROM SERVER")
            for rows in range(1,sheet.nrows):
                w_sheet.cell(row=rows+1,column=sheet.ncols+1).value=cliet.service.GetMinClassRating(sheet.cell(rows,data_column['TEMPERATURE']).value,sheet.cell(rows,data_column['PRESSURE']).value,sheet.cell(rows,data_column['VALVE MATERIAL']).value)
                
        work.save(os.getcwd()+'/'+'Output.xls')
        order=[]
        ssd=standard_list.sheet_by_name('standard')
        for col in range(ssd.ncols):
            order.append(ssd.cell(0,col).value)

        final_s=work.create_sheet('final')
        regret_s=work.create_sheet('regret')
        wb2=open_workbook(os.getcwd()+'/'+'Output.xls' )
        sheet=wb2.sheet_by_index(0)
        counter=1
        counter2=1
        head=1
        redFill = PatternFill(start_color='00CED1FF',
                    end_color='00CED1FF',
                    fill_type='solid')
        for row in range(sheet.nrows):
            count=0

            while count < len(order):
                if head==1:
                    for col in range(len(order)):
                        if order[col] in data_column:
                            final_s.cell(row=counter,column=col+3).value=order[col]
                            final_s.cell(row=counter,column=2).value='VALVE PROJECT'
                            final_s.cell(row=counter,column=1).value='VALVE SCHEDULE'
                    count+=1
                    counter+=1
                    head=0
                elif sheet.cell(row,data_column[order[count]]).value in alias_data['VALVE TYPE'].keys():
                    for col in range(len(order)):
                        if order[col] in data_column :
                            if w_sheet.cell(row=row+1,column=data_column[order[col]]+1).fill==redFill:
                                final_s.cell(row=counter,column=col+3).fill=redFill
                                final_s.cell(row=counter,column=col+3).value=sheet.cell(row,data_column[order[col]]).value
                                final_s.cell(row=counter,column=2).value=self.vp
                                final_s.cell(row=counter,column=1).value=self.vs

                            else:
                                final_s.cell(row=counter,column=col+3).value=sheet.cell(row,data_column[order[col]]).value
                                final_s.cell(row=counter,column=2).value=self.vp
                                final_s.cell(row=counter,column=1).value=self.vs
                    counter+=1
                    count+=1
                    break
                else:
                    for col in range(len(order)):
                        if order[col] in data_column :
                            if w_sheet.cell(row=row+1,column=data_column[order[col]]+1).fill==redFill:
                                regret_s.cell(row=counter2,column=col+3).fill=redFill
                                regret_s.cell(row=counter2,column=col+3).value=sheet.cell(row,data_column[order[col]]).value
                                regret_s.cell(row=counter2,column=2).value=self.vp
                                regret_s.cell(row=counter2,column=1).value=self.vs
                            else:
                                regret_s.cell(row=counter2,column=col+3).value=sheet.cell(row,data_column[order[col]]).value
                                regret_s.cell(row=counter2,column=2).value=self.vp
                                regret_s.cell(row=counter2,column=1).value=self.vs
                    counter2+=1
                    count+=1
                    break
        work.save(os.getcwd()+'/'+'Output.xls')
        wb2=open_workbook(os.getcwd()+'/'+'Output.xls' )
        Part_number_extraction(alias_data2,standard_list,wb2,work,self.size_value,convert_i_2_nb,self.part_value,self.part_value2,convert_nb_2_i)
        work.save(os.getcwd()+'/'+'Output.xls')
