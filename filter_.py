import os
from openpyxl.styles import Color, PatternFill, Font, Border
from xlrd import *
from openpyxl import Workbook,load_workbook
from extraction import Extraction
class Filter(Extraction):
    def non_zero_row(self,test_row,col_names,q_alias):
        non_zero=0
        check=False
        check3=False
        nil=0
        flag=0
        q_col=0
        number=''
        for i in col_names:
            try:
                if Extraction.capitalised(self,i[0]) in q_alias :
                    q_col=i[1]
                    break
                else:
                    for temp_val in q_alias:
                        if temp_val in Extraction.capitalised(self,i[0]):
                            q_col=i[1]
                            break
            except Exception as e:
                pass

        for i in range(len(test_row)):
                if test_row[i]=='':
                   nil+=1
        
        if test_row[q_col]!='':
            if type(test_row[q_col])==int or type(test_row[q_col])==float:
                check3=True
        if check3:
            if len(test_row)-2*nil>=0:
                return True

            else:
                return True

        else:
            return False

    def verify_material(self,conversion,size_type,key,temp,size,a105,f22,f91,f92,wcb,wcc,wc9,c12a,c12ab):
        if size_type=='IN':
            size=str(size).replace('\\','/').split('-')
        else:
            size=str(conversion(float(size)))[:-2].replace('\\','/').split('-')
        if '/' in str(size) and len(size)>1:
            a,b=size
            b=b.replace('/','')
            size=int(a)+float(b[0])/float(b[1])
        elif '/' in str(size):
            b,=size
            b=b.replace('/','')
            size=float(b[0])/float(b[1])
        elif 'X' not in str(size) and '' not in size:
            b,=size
            size=float(b)
        if key in ['NRV','PLNRV','SCNRV']:
            if size<=2:
                if key =='PLNRV':
                    return True , key
                else:
                    return False ,'PLNRV'
            if size>2:
                if key =='SCNRV':
                    return True , key
                else:
                    return False ,'SCNRV'
            
        if temp!='':
            if temp<=a105:
                if size<=2 and key=='A105':
                    return True,key
                elif size<=2:
                    return False,'A105'
            elif temp<=wcb:
                if size>2 and key=='WCB':
                    return True,key
                elif size>2:
                    return False,'WCB'
            elif temp<=f22 :
                if size<=2 and key=='F22':
                    return True,key
                elif size<=2:
                        return False,'F22'
            elif temp<=wc9:
                if size>2 and key=='WC9':
                    return True,key
                elif size>2:
                    return False,'WC9'
            elif temp<=f91 :
                if size<=2 and key=='F91':
                    return True,key
                elif size<=2:
                    return False,'F91'
            elif temp<=c12a:
                if size>2 and key=='C12A':
                    return True,key
                elif size>2:
                    return False,'C12A'

            elif temp<=f92 :
                if size<=2 and key=='F92':
                    return True,key
                elif size<=2:
                    return False,'F92'
            elif temp<c12ab:
                if size>2 and key=='C12A':
                    return True,key
                elif size>2:
                    return False,'C12A'
            else:
                return True,key
        else:
            return False,key
    def formatting_data(self,convert_nb_2_i,size_type,source_wb,dest_wb,data_col,alias_data,material_check,a105,f22,f91,f92,wcb,wcc,wc9,c12a,c12ab):
        sheet=source_wb.sheet_by_name('Standard Sheet')
        sheet_f=dest_wb.get_sheet_by_name('Standard Sheet')
        redFill = PatternFill(start_color='00CED1FF',
                end_color='00CED1FF',
                fill_type='solid')
        for data in alias_data:
            print "FORMATTING FOR -"+str(data)
            for row_s in range(sheet.nrows):
                try:
                    cap=sheet.cell(row_s,data_col[data]).value
                except Exception as e:
                    break
                if '-' not in str(cap)[:1]:
                    cap=Extraction.capitalised(self,str(cap))
                for key in alias_data[data]:
                    if cap in alias_data[data][key] :
                        sheet_f.cell(row=row_s+1,column=data_col[data]+1).value=key
                        if data=='VALVE MATERIAL' and material_check=='YES' and key not in ['F316','CF3','CF8']:
                                try:
                                    check,material=self.verify_material(convert_nb_2_i,size_type,key,sheet.cell(row_s,data_col['TEMPERATURE']).value,sheet.cell(row_s,data_col['VALVE SIZE']).value,a105,f22,f91,f92,wcb,wcc,wc9,c12a,c12ab)
                                    if check:
                                        sheet_f.cell(row=row_s+1,column=data_col[data]+1).value=material
                                    else:
                                        sheet_f.cell(row=row_s+1,column=data_col[data]+1).value=material
                                        sheet_f.cell(row=row_s+1,column=data_col[data]+1).fill=redFill
                                except Exception as e :
                                   print e
                                   pass
                        break
        dest_wb.save(os.getcwd()+'/'+'Output.xls')
